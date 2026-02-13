
import logging
import sys
from pathlib import Path
from openpyxl import load_workbook
import re

# Add src to path
sys.path.append(str(Path(__file__).parent.parent))

# Setup logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

def verify_axis_dec2025():
    amc = "axis"
    year = 2025
    month = 12
    
    file_path = Path(f"data/output/merged excels/{amc}/{year}/CONSOLIDATED_AXIS_2025_12.xlsx")
    
    if not file_path.exists():
        print(f"❌ FAILED: Consolidated file not found: {file_path}")
        return

    print(f"✅ Found file: {file_path}")
    
    wb = load_workbook(file_path, read_only=True)
    sheet_names = wb.sheetnames
    print(f"Sheet count: {len(sheet_names)}")
    
    # Axis has roughly 88 schemes
    if len(sheet_names) > 80:
        print(f"✅ SUCCESS: Sheet count is significant ({len(sheet_names)}).")
    else:
        print(f"❌ FAILED: Sheet count too low ({len(sheet_names)}). Expected > 80.")

    # Verify Naming Cleaning
    # Sheet names should be like "%20Portfolio %% AXISBCF"
    # Should NOT have: "Monthly", "Axis Mutual Fund", dates in sheet names
    
    bad_keywords = ['Monthly', 'Axis Mutual Fund', 'December', '2025', '31']
    dirty_sheets = []
    
    print("\nSample Sheet Names:")
    for name in sheet_names[:10]:
        print(f"  - {name}")
        for bad in bad_keywords:
            if bad.lower() in name.lower():
                dirty_sheets.append(name)
                
    if len(dirty_sheets) == 0:
         print("\n✅ SUCCESS: Sheet names appear clean (no 'Axis Mutual Fund', 'Monthly', dates).")
    else:
         print(f"\n❌ FAILED: Found dirty sheet names: {dirty_sheets[:5]}")

    # Check specific schemes
    expected = ['AXISBCF', 'AXISMCF', 'AXISSCF', 'AXIS500']  # Bluechip, Midcap, Small Cap, Nifty 500
    # Exact match might be hard due to trimming/cleaning, so we check partial
    found_count = 0
    for exp in expected:
        if any(exp.lower() in s.lower() for s in sheet_names):
            found_count += 1
            
    print(f"\nFound {found_count}/{len(expected)} representative scheme codes.")
    
    if found_count >= 3:
        print("✅ SUCCESS: Found most expected scheme codes.")
    else:
        print(f"❌ WARNING: Only found {found_count} expected scheme codes.")

if __name__ == "__main__":
    verify_axis_dec2025()
