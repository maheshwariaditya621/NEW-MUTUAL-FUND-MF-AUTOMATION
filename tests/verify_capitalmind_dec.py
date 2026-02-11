
import logging
import sys
from pathlib import Path
from openpyxl import load_workbook

# Add src to path
sys.path.append(str(Path(__file__).parent.parent))

# Setup logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

def verify_capitalmind_dec2025():
    amc = "capitalmind"
    year = 2025
    month = 12
    
    file_path = Path(f"data/output/merged excels/{amc}/{year}/CONSOLIDATED_CAPITALMIND_2025_12.xlsx")
    
    if not file_path.exists():
        print(f"❌ FAILED: Consolidated file not found: {file_path}")
        return

    print(f"✅ Found file: {file_path}")
    
    wb = load_workbook(file_path, read_only=True)
    sheet_names = wb.sheetnames
    print(f"Sheet count: {len(sheet_names)}")
    print(f"Sheet names: {sheet_names}")
    
    # Capitalmind has a few schemes
    if len(sheet_names) >= 2:
        print(f"✅ SUCCESS: Sheet count is valid ({len(sheet_names)}).")
    else:
        print(f"❌ FAILED: Sheet count too low ({len(sheet_names)}).")
        
    # Check for clean names
    expected = ['CMFLEXI', 'CMLIQ']
    found = [s for s in sheet_names if s in expected]
    if len(found) == len(expected):
        print(f"✅ SUCCESS: Found clean expected sheet names: {found}")
    else:
        print(f"❌ FAILED: Sheet names might not be clean enough. Found: {sheet_names}")

if __name__ == "__main__":
    verify_capitalmind_dec2025()
