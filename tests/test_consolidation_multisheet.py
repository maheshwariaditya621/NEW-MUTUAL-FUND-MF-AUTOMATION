
import logging
import sys
import shutil
from pathlib import Path
from openpyxl import Workbook, load_workbook

# Add src to path
sys.path.append(str(Path(__file__).parent.parent))

from src.utils.excel_merger import consolidate_amc_downloads

# Setup logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

def create_excel(path: Path, sheets: dict):
    """Create a dummy excel file with specified sheets."""
    wb = Workbook()
    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])
        
    for name, content in sheets.items():
        ws = wb.create_sheet(title=name)
        ws["A1"] = content
        
    wb.save(path)

def test_multisheet_consolidation():
    amc = "test_amc_multisheet"
    year = 2025
    month = 12
    
    # Setup paths
    raw_folder = Path(f"data/raw/{amc}/{year}_{month:02d}")
    if raw_folder.exists():
        shutil.rmtree(raw_folder)
    raw_folder.mkdir(parents=True, exist_ok=True)
    
    output_folder = Path(f"data/output/merged excels/{amc}/{year}")
    if output_folder.exists():
        shutil.rmtree(output_folder)
    
    print("\n--- TEST: Multi-Sheet Consolidation ---")
    
    # Case A: Single Sheet (Should use Filename)
    # File: Single_Sheet_File.xlsx -> Sheet: "Single Sheet File"
    create_excel(raw_folder / "Single_Sheet_File.xlsx", {"Sheet1": "Data1"})
    
    # Case B: Multi Sheet (Should use Sheet Titles)
    # File: Multi_Sheet_File.xlsx -> Sheets: "Scheme Alpha", "Scheme Beta"
    create_excel(raw_folder / "Multi_Sheet_File.xlsx", {
        "Scheme Alpha From AMC": "Data A", 
        "Scheme Beta 2024": "Data B"
    })
    
    logger.info(f"Created dummy files in {raw_folder}")
    
    # Run consolidation
    result = consolidate_amc_downloads(amc, year, month)
    
    # Verification
    if not result or not result.exists():
        print("❌ FAILED: Consolidated file NOT created.")
        return

    print(f"✅ Consolidated file created: {result.name}")
    
    wb = load_workbook(result)
    sheet_names = wb.sheetnames
    print(f"DEBUG: Found sheets: {sheet_names}")
    
    # Assertions
    # 1. Single Sheet File Check
    # Filename: "Single_Sheet_File.xlsx" -> Cleaned: "Single Sheet File"
    assert "Single Sheet File" in sheet_names, f"Missing 'Single Sheet File'. Found: {sheet_names}"
    
    # 2. Multi Sheet File Check
    # Sheet: "Scheme Alpha From AMC" -> Cleaned: "Scheme Alpha From" (AMC removed)
    # Sheet: "Scheme Beta 2024" -> Cleaned: "Scheme Beta" (2024 removed)
    if "Scheme Alpha From" in sheet_names:
        print("✅ Found 'Scheme Alpha From'")
    else:
        print(f"❌ Missing 'Scheme Alpha From'. Found: {sheet_names}")
        return

    if "Scheme Beta" in sheet_names:
        print("✅ Found 'Scheme Beta'")
    else:
        print(f"❌ Missing 'Scheme Beta'. Found: {sheet_names}")
        return
    
    print("✅ SUCCESS: Sheet naming logic verified.")
    print("  - Single-sheet files used FILENAME.")
    print("  - Multi-sheet files used SHEET TITLES.")
    print("  - Cleaning logic applied correctly.")

    print("\n🎉 MULTI-SHEET CONSOLIDATION TEST PASSED!")

    # Cleanup
    if raw_folder.exists():
        shutil.rmtree(f"data/raw/{amc}")
    if output_folder.exists():
        shutil.rmtree(f"data/output/merged excels/{amc}")

if __name__ == "__main__":
    test_multisheet_consolidation()
