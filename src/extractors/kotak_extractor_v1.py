import pandas as pd
import re
from typing import List, Dict, Any, Optional
from datetime import datetime
from openpyxl import load_workbook

from src.extractors.base_extractor import BaseExtractor
from src.config.constants import AMC_KOTAK
from src.config import logger

class KotakExtractorV1(BaseExtractor):
    """
    Extractor for Kotak Mutual Fund (Dec 2025 format).
    
    Structure:
    - Scheme Name: Row 0 ("Portfolio of ... as on ...")
    - Header: Row 1
    - Columns: 
        - Name of Instrument
        - ISIN Code
        - Market Value (Rs.in Lacs) -> Needs * 100,000
        - % to Net Assets -> Percentage
    """
    
    def __init__(self):
        super().__init__(amc_name=AMC_KOTAK, version="v1")
        
        # Column mapping based on analysis
        self.column_mapping = {
            "Name of Instrument": "security_name",
            "Name of the Instrument": "security_name", # Possible variant
            "ISIN Code": "isin",
            "ISIN": "isin",
            "Industry": "sector",
            "Quantity": "quantity",
            "Market Value (Rs.in Lacs)": "market_value",
            "Market Value(Rs. in Lakhs)": "market_value", # Possible variant
            "% to Net Assets": "percent_of_nav"
        }

    def standardize_columns(self, df: pd.DataFrame, mapping: Dict[str, str]) -> pd.DataFrame:
        """
        Standardize column names based on mapping.
        """
        new_cols = {}
        for col in df.columns:
            col_str = str(col).strip()
            # Fuzzy match for complex headers
            for key, val in mapping.items():
                if key in col_str:
                    new_cols[col] = val
                    break
        return df.rename(columns=new_cols)

    def _extract_scheme_name(self, df_raw: pd.DataFrame, sheet_name: str) -> str:
        """
        Extract scheme name from Row 0 (Cell A1/B1/C1).
        Pattern: "Portfolio of [SCHEME NAME] as on [DATE]"
        """
        # Iterate first few cells in first row to find the text
        for col in range(5):
            val = str(df_raw.iloc[0, col])
            if "Portfolio of " in val:
                # Regex to extract name
                match = re.search(r"Portfolio of\s+(.+?)\s+as on", val, re.IGNORECASE)
                if match:
                    return match.group(1).strip()
        
        # Fallback to sheet name
        return sheet_name

    def extract(self, file_path: str) -> List[Dict[str, Any]]:
        """
        Delegates to multi-scheme extraction logic to handle stacked schemes (e.g. Midcap 150).
        """
        logger.info(f"[Kotak] Starting extraction from: {file_path}")
        return self._extract_multi_scheme_sheet(file_path)

    def _extract_multi_scheme_sheet(self, file_path: str) -> List[Dict[str, Any]]:
        wb = load_workbook(file_path, read_only=True, data_only=True)
        all_holdings = []
        schemes_processed = 0
        schemes_with_equity = 0
        
        for sheet_name in wb.sheetnames:
            try:
                if self._should_skip_sheet(sheet_name):
                    continue
                    
                df_raw = pd.read_excel(file_path, sheet_name=sheet_name, header=None)
                
                # Find all scheme start rows
                scheme_starts = []
                for idx, row in df_raw.iterrows():
                    # Check first 5 cols for "Portfolio of"
                    row_str = " ".join([str(x) for x in row.iloc[:5] if pd.notna(x)])
                    if "Portfolio of " in row_str and " as on " in row_str:
                         scheme_starts.append(idx)
                
                if not scheme_starts:
                    # Fallback: Maybe just 1 scheme without "Portfolio of"? 
                    # Try finding header directly.
                    scheme_starts = [0]
                    
                # Process each block
                for i, start_row in enumerate(scheme_starts):
                    end_row = scheme_starts[i+1] if i + 1 < len(scheme_starts) else len(df_raw)
                    
                    # Slice block
                    df_block = df_raw.iloc[start_row:end_row].copy().reset_index(drop=True)
                    
                    if df_block.empty: continue
                    
                    # A. Extract Name (Row 0 of block)
                    raw_name_cell = " ".join([str(x) for x in df_block.iloc[0].values if pd.notna(x)])
                    match = re.search(r"Portfolio of\s+(.+?)\s+as on", raw_name_cell, re.IGNORECASE)
                    if match:
                        raw_scheme_name = match.group(1).strip()
                    else:
                        # Fallback if we guessed start_row=0
                        raw_scheme_name = sheet_name
                    
                    scheme_name_clean = self._clean_scheme_name(raw_scheme_name)
                    
                    try:
                        scheme_info = self.parse_verbose_scheme_name(scheme_name_clean)
                    except:
                        continue
                        
                    # B. Find Header in Block
                    header_keywords = ["Instrument", "ISIN", "Quantity", "Market Value"]
                    header_idx = self.find_header_row(df_block, header_keywords)
                    
                    if header_idx == -1:
                        continue
                        
                    # C. Extract Data
                    # Handle merged cells via ffill on header? 
                    # df_block is from read_excel(header=None), so no merged cell awareness from pandas directly 
                    # unless we re-read? No, read_excel handles merge by placing value in top-left.
                    # We need ffill on the header row itself.
                    
                    header_row_vals = df_block.iloc[header_idx].ffill()
                    df_block.columns = header_row_vals
                    
                    df_data = df_block.iloc[header_idx+1:].reset_index(drop=True)
                    
                    # Standardize
                    df_clean = self.standardize_columns(df_data, self.column_mapping)
                    
                    if 'isin' not in df_clean.columns: continue

                    # Filter Equity
                    try:
                        df_equity = self.filter_equity_isins(df_clean, 'isin')
                    except:
                        continue
                        
                    if df_equity.empty: 
                        schemes_processed += 1
                        continue
                        
                    # D. Extract Total Net Assets (AUM) from the block's footer
                    raw_net_assets = None
                    for idx, row in df_clean.iterrows():
                        row_vals = [str(val).upper() if pd.notna(val) else "" for val in row.values]
                        row_text = " ".join(row_vals)
                        if "GRAND TOTAL" in row_text or "NET ASSETS" in row_text or "TOTAL AUM" in row_text:
                            candidates = []
                            for val in row.values:
                                f_val = self.safe_float(val)
                                if f_val is not None and f_val > 105: # Avoid 100.00%
                                    candidates.append(f_val)
                            
                            if candidates:
                                if len(candidates) > 1:
                                     if abs(candidates[-1] - 100.0) < 0.05:
                                         raw_net_assets = candidates[-2]
                                     else:
                                         raw_net_assets = candidates[-1]
                                else:
                                    raw_net_assets = candidates[0]
                                
                            if raw_net_assets:
                                break
                                
                    normalized_net_assets = None
                    if raw_net_assets:
                        normalized_net_assets = self.normalize_currency(raw_net_assets, "LAKHS")

                    block_holdings = []
                    if not df_equity.empty:
                        for _, row in df_equity.iterrows():
                            sec_name = str(self._resolve_merged_column(row, 'security_name') or '').lower()
                            
                            # Guard: Skip Total/Sub Total rows if they pass ISIN filter, but DO NOT BREAK (as multiple schemes might follow)
                            if 'total' in sec_name or 'sub total' in sec_name or 'net current' in sec_name:
                                continue
                                
                            isin = self._resolve_merged_column(row, 'isin')
                            qty = self.safe_float(self._resolve_merged_column(row, 'quantity'))
                            
                            try:
                                mv_val = self._resolve_merged_column(row, 'market_value')
                                mv_inr = self.safe_float(mv_val) * 100000
                            except: mv_inr = 0.0
                            
                            try:
                                nav_val = self._resolve_merged_column(row, 'percent_of_nav')
                                nav_pct = self.safe_float(nav_val)
                            except: nav_pct = 0.0
                            
                            if qty < 0 or mv_inr < 0: continue
                            
                            holding = {
                                "amc_name": self.amc_name,
                                "scheme_name": scheme_info['scheme_name'],
                                "scheme_description": scheme_info['description'],
                                "plan_type": scheme_info['plan_type'],
                                "option_type": scheme_info['option_type'],
                                "is_reinvest": scheme_info['is_reinvest'],
                                "isin": isin,
                                "company_name": self.clean_company_name(sec_name),
                                "quantity": int(qty),
                                "market_value_inr": mv_inr,
                                "percent_of_nav": nav_pct,
                                "sector": str(self._resolve_merged_column(row, 'sector') or '').strip() or None,
                                "total_net_assets": normalized_net_assets
                            }
                            block_holdings.append(holding)
                    elif normalized_net_assets:
                        # Ghost Holding for Non-Equity funds
                        holding = {
                            "amc_name": self.amc_name,
                            "scheme_name": scheme_info['scheme_name'],
                            "scheme_description": scheme_info['description'],
                            "plan_type": scheme_info['plan_type'],
                            "option_type": scheme_info['option_type'],
                            "is_reinvest": scheme_info['is_reinvest'],
                            "isin": None,
                            "company_name": "N/A",
                            "quantity": 0,
                            "market_value_inr": 0,
                            "percent_of_nav": 0,
                            "sector": "N/A",
                            "total_net_assets": normalized_net_assets
                        }
                        block_holdings.append(holding)
                        
                    self.validate_nav_completeness(block_holdings, scheme_info['scheme_name'])
                    if block_holdings:
                        all_holdings.extend(block_holdings)
                        schemes_with_equity += 1
                        logger.info(f"[Kotak] ✓ {scheme_info['scheme_name']}: {len(block_holdings)} holdings (Block)")
                        
                    schemes_processed += 1
                    
            except Exception as e:
                logger.error(f"[Kotak] Error processing sheet {sheet_name}: {e}")
                
        logger.info(f"[Kotak] Extraction complete. Found {len(all_holdings)} holdings.")
        return all_holdings


    def _clean_scheme_name(self, raw_name: str) -> str:
        """
        Clean and normalize scheme name to ensure it starts with 'KOTAK MAHINDRA '.
        """
        cleaned = raw_name.strip().upper()
        
        # Remove "KOTAK MAHINDRA MUTUAL FUND -" prefix first to avoid double prefixing
        if "KOTAK MAHINDRA MUTUAL FUND" in cleaned:
             cleaned = cleaned.replace("KOTAK MAHINDRA MUTUAL FUND", "").strip()
             if cleaned.startswith("-"):
                 cleaned = cleaned[1:].strip()

        # Ensure it starts with KOTAK MAHINDRA
        base_prefix = "KOTAK MAHINDRA "
        
        if cleaned.startswith(base_prefix):
            return cleaned
            
        elif cleaned.startswith("KOTAK MAHINDRA"): # Missing space
             return base_prefix + cleaned[14:].strip()
             
        elif cleaned.startswith("KOTAK "):
             # e.g. "KOTAK FLEXICAP FUND" -> "KOTAK MAHINDRA FLEXICAP FUND"
             return base_prefix + cleaned[5:].strip()
             
        else:
             # Prefix missing entirely, e.g. "AGGRESSIVE HYBRID FUND"
             return base_prefix + cleaned

    def _should_skip_sheet(self, sheet_name: str) -> bool:
        """Skip summary/index sheets."""
        skip_terms = ['summary', 'index', 'contents', 'disclaimer', 'note']
        sn = sheet_name.lower()
        return any(term in sn for term in skip_terms)

    def _resolve_merged_column(self, row: pd.Series, col_name: str) -> Any:
        """
        Resolve value from a column that might be duplicated due to header ffill.
        Returns the first non-empty value found.
        """
        val = row.get(col_name)
        if isinstance(val, pd.Series):
            # Multiple columns with same name.
            # Look for first non-empty value.
            for v in val.values:
                if pd.notna(v) and str(v).strip() != '':
                    return v
            return None
        return val
