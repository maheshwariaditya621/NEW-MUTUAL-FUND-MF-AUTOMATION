import os
import glob
import re
import shutil
import zipfile
import subprocess
from copy import copy
from pathlib import Path
from typing import List, Optional
from urllib.parse import unquote
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE

from src.config import logger
from src.alerts.telegram_notifier import get_notifier

def copy_style(src_cell, dest_cell):
    """Deep copies style attributes from source cell to destination cell."""
    if src_cell.has_style:
        dest_cell.font = copy(src_cell.font)
        dest_cell.border = copy(src_cell.border)
        dest_cell.fill = copy(src_cell.fill)
        dest_cell.number_format = copy(src_cell.number_format)
        dest_cell.protection = copy(src_cell.protection)
        dest_cell.alignment = copy(src_cell.alignment)

def merge_project_excels_openpyxl(folder_path: Path, output_filename: Path) -> Optional[Path]:
    """
    Optimized version with style caching to handle large volumes of files.
    """
    logger.info(f"Starting OpenPyXL Merge in: {folder_path}")
    
    _prepare_raw_folder(folder_path)

    all_files = glob.glob(str(folder_path / "*.xlsx"))
    files = sorted([f for f in all_files if not os.path.basename(f).startswith("~$")])
    
    if not files:
        logger.warning(f"No Excel files found in {folder_path} to merge.")
        return None

    master_wb = Workbook()
    if "Sheet" in master_wb.sheetnames:
        master_wb.remove(master_wb["Sheet"])
    
    for idx, file_path in enumerate(files):
        # Decode filename to handle %20 and other URL-encoded characters
        file_name = unquote(os.path.basename(file_path))
        
        # Avoid merging an already generated consolidation file if it happens to be in the same folder
        if file_name == output_filename.name:
            continue

        try:
            # CRITICAL: Use read_only=True for large files (e.g. HDFC 1M row sheets)
            # This avoids memory exhaustion and hangs.
            src_wb = load_workbook(file_path, data_only=True, read_only=True)
            worksheets = src_wb.worksheets
            
            logger.info(f"[{idx+1}/{len(files)}] Processing {file_name} ({len(worksheets)} sheets)")

            for src_ws in worksheets:
                # --- Naming Logic ---
                if len(worksheets) == 1:
                    base_raw_name = file_name.rsplit('.', 1)[0]
                else:
                    # Combine filename and sheet title to ensure uniqueness and context
                    fname_clean = file_name.rsplit('.', 1)[0]
                    base_raw_name = f"{fname_clean} {src_ws.title}"
                
                # --- Cleaning Logic ---
                # 1. First, normalize all delimiters to spaces so regex \b works
                sheet_name = base_raw_name.replace("_", " ").replace("-", " ").strip()
                
                # Keywords to strip
                days = [str(i).zfill(2) for i in range(1, 40)] + [str(i) for i in range(1, 40)]
                strip_keywords = [
                    'ICICI', 'Prudential', 'HDFC', 'SBI', 'Axis', 'Nippon', 'India', 'Tata', 'Quant', 
                    'Bandhan', 'Kotak', 'Aditya', 'Birla', 'Sun', 'Life', 'Mutual', 'Fund', 'MF', 'AMC',
                    'Canara', 'Robeco', 'BNP', 'Paribas', 'Capitalmind',
                    'Monthly', 'Portfolio', 'Disclosure', 'Statement',
                    'January', 'February', 'March', 'April', 'May', 'June', 'July', 'August', 
                    'September', 'October', 'November', 'December', '2024', '2025', '2026',
                    'Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec',
                ] + days
                
                # 2. Strip keywords (with word boundaries)
                pattern = r'\b(' + '|'.join(set(strip_keywords)) + r')\b'
                sheet_name = re.sub(pattern, '', sheet_name, flags=re.IGNORECASE)
                
                # 3. Strip hex-like strings (4+ chars) but PROTECT years (4 digits)
                sheet_name = re.sub(r'\b[0-9a-f]{4,}\b', '', sheet_name, flags=re.IGNORECASE)
                
                # 4. Final Cleanup: Resize multiple spaces to single
                sheet_name = re.sub(r'\s+', ' ', sheet_name).strip()
                
                # Final safety check for empty name
                if not sheet_name or len(sheet_name) < 3:
                    sheet_name = base_raw_name[:31]
                    
                # Limit to 31 chars (Excel limit)
                base_sheet_name = sheet_name[:31].strip()
                sheet_name = base_sheet_name
                
                # Handle Duplicates
                counter = 1
                while sheet_name in master_wb.sheetnames:
                    suffix = f" ({counter})"
                    available_len = 31 - len(suffix)
                    sheet_name = f"{base_sheet_name[:available_len].strip()}{suffix}"
                    counter += 1

                logger.debug(f"    Merging sheet: '{src_ws.title}' -> '{sheet_name}'")
                
                dest_ws = master_wb.create_sheet(title=sheet_name)

                # In read_only mode, we can't always trust max_row, so we rely more on the loop
                max_r = src_ws.max_row
                max_c = src_ws.max_column
                
                found_first_data = False
                consecutive_empty_rows = 0
                MAX_EMPTY_ROWS = 50  # Stop after 50 empty rows AFTER data starts
                
                for r_idx, row in enumerate(src_ws.iter_rows(max_row=max_r, max_col=max_c), 1):
                    # Check if row is empty
                    is_empty_row = all(cell.value is None for cell in row)
                    
                    if is_empty_row:
                        if found_first_data:
                            consecutive_empty_rows += 1
                            if consecutive_empty_rows > MAX_EMPTY_ROWS:
                                logger.debug(f"    Stopping at row {r_idx} due to {MAX_EMPTY_ROWS} consecutive empty rows.")
                                break
                        continue
                    else:
                        found_first_data = True
                        consecutive_empty_rows = 0
                    
                    # Copy row height specific to row (skip in read-only)
                    if not src_wb.read_only and r_idx in src_ws.row_dimensions:
                        if src_ws.row_dimensions[r_idx].customHeight:
                            dest_ws.row_dimensions[r_idx].height = src_ws.row_dimensions[r_idx].height

                    for c_idx, src_cell in enumerate(row, 1):
                        val = src_cell.value
                        if val is None:
                            continue
                        
                        dest_cell = dest_ws.cell(row=r_idx, column=c_idx, value=val)
                        
                        if src_cell.has_style:
                            dest_cell.number_format = src_cell.number_format
                            if r_idx < 20: # Conserve headers/top rows style
                                dest_cell.font = copy(src_cell.font)
                                dest_cell.border = copy(src_cell.border)
                                dest_cell.fill = copy(src_cell.fill)
                                dest_cell.alignment = copy(src_cell.alignment)

                if not src_wb.read_only:
                    for col_letter, col_dim in src_ws.column_dimensions.items():
                        if col_dim.customWidth:
                            dest_ws.column_dimensions[col_letter].width = col_dim.width

                if not src_wb.read_only and hasattr(src_ws, 'merged_cells') and src_ws.merged_cells:
                    for merged_range in src_ws.merged_cells.ranges:
                        try:
                            dest_ws.merge_cells(str(merged_range))
                        except:
                            pass 

            src_wb.close()

        except Exception as e:
            logger.error(f"Error processing file {file_name}: {e}")

    if len(master_wb.sheetnames) == 0:
        logger.warning(f"No valid sheets were created for {output_filename}")
        return None

    output_filename.parent.mkdir(parents=True, exist_ok=True)
    master_wb.save(str(output_filename))
    logger.info(f"Consolidated file saved: {output_filename}")
    return output_filename

def merge_project_excels_com(folder_path: Path, output_filename: Path) -> Optional[Path]:
    """
    High-fidelity merge using Microsoft Excel via COM.
    Preserves all formatting, merged cells, and column widths.
    Required for AMCs like HDFC with large sheets and complex formatting.
    """
    try:
        import win32com.client
        import pythoncom
    except ImportError:
        logger.warning("pywin32 not installed. COM merge unavailable.")
        return None

    logger.info(f"Starting High-Fidelity COM Merge in: {folder_path}")
    
    _prepare_raw_folder(folder_path)

    files = [f for f in glob.glob(str(folder_path / "*.xlsx")) if not os.path.basename(f).startswith("~$")]
    if not files:
        logger.warning(f"No Excel files found in {folder_path} to merge.")
        return None
    files.sort()

    pythoncom.CoInitialize()
    excel = None
    try:
        # Use DispatchEx to ensure a fresh, separate instance
        excel = win32com.client.DispatchEx("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        excel.ScreenUpdating = False

        # 0. Extraction and Conversion - USE THE SAME excel instance
        _prepare_raw_folder(folder_path, excel_app=excel)

        files = [f for f in glob.glob(str(folder_path / "*.xlsx")) if not os.path.basename(f).startswith("~$")]
        if not files:
            logger.warning(f"No Excel files found in {folder_path} to merge.")
            return None
        files.sort()

        master_wb = excel.Workbooks.Add()
        # Remove default sheets if any
        while master_wb.Sheets.Count > 1:
            master_wb.Sheets(1).Delete()
        
        dummy_sheet = master_wb.Sheets(1)
        dummy_sheet.Name = "TEMP_INITIAL_SHEET"

        for idx, file_path in enumerate(files):
            file_name = unquote(os.path.basename(file_path))
            if file_name == output_filename.name:
                continue

            logger.info(f"[{idx+1}/{len(files)}] COM Processing {file_name}")
            abs_path = str(Path(file_path).resolve())
            
            try:
                src_wb = excel.Workbooks.Open(abs_path, ReadOnly=True, UpdateLinks=False)
                
                for s_idx in range(1, src_wb.Sheets.Count + 1):
                    src_sheet = src_wb.Sheets(s_idx)
                    clean_name = _get_clean_sheet_name(file_name, src_sheet.Name, [s.Name for s in master_wb.Sheets])
                    
                    target_after = master_wb.Sheets(master_wb.Sheets.Count)
                    src_sheet.Copy(None, target_after)
                    
                    new_sheet = master_wb.Sheets(master_wb.Sheets.Count)
                    try:
                        new_sheet.Name = clean_name
                    except:
                        safe_name = re.sub(r'[*?:\/\[\]]', '', clean_name)[:31]
                        try:
                            new_sheet.Name = safe_name
                        except:
                            new_sheet.Name = f"Sheet_{master_wb.Sheets.Count}"

                src_wb.Close(False)
            except Exception as e:
                logger.error(f"  COM failed to process {file_name}: {e}")

        # Remove the dummy sheet
        try:
            excel.DisplayAlerts = False
            master_wb.Sheets("TEMP_INITIAL_SHEET").Delete()
        except:
            pass

        output_filename.parent.mkdir(parents=True, exist_ok=True)
        abs_output = str(output_filename.resolve())
        master_wb.SaveAs(abs_output, FileFormat=51)
        master_wb.Close()
        
        logger.info(f"Consolidated file saved (COM): {output_filename}")
        return output_filename

    except Exception as e:
        logger.error(f"Global COM Error: {e}")
        return None
    finally:
        if excel:
            try: 
                excel.DisplayAlerts = False
                excel.Quit()
            except: pass
        pythoncom.CoUninitialize()

def _prepare_raw_folder(folder_path: Path, excel_app=None):
    """Extraction and Conversion logic."""
    # 0. Pre-processing: Extract ZIPs
    zip_files = glob.glob(str(folder_path / "*.zip"))
    for zip_path_str in zip_files:
        zip_path = Path(zip_path_str)
        try:
            with zipfile.ZipFile(zip_path, 'r') as zip_ref:
                excel_files_in_zip = [f for f in zip_ref.namelist() if f.lower().endswith(('.xlsx', '.xls'))]
                if excel_files_in_zip:
                    logger.info(f"Extracting Excel files from {zip_path.name}...")
                    for f in excel_files_in_zip: zip_ref.extract(f, folder_path)
        except Exception as e: logger.error(f"ZIP error: {e}")
    
    _convert_xls_to_xlsx(folder_path, excel_app)

def is_soffice_available() -> bool:
    """Checks if LibreOffice 'soffice' command is available in the system path."""
    return shutil.which('soffice') is not None

def convert_xls_to_xlsx_soffice(xls_path: Path, xlsx_path: Path) -> bool:
    """
    Converts .xls to .xlsx using LibreOffice (soffice) headless command.
    High-fidelity solution for Linux/macOS.
    """
    try:
        out_dir = str(xlsx_path.parent.resolve())
        cmd = [
            'soffice', 
            '--headless', 
            '--convert-to', 'xlsx', 
            '--outdir', out_dir, 
            str(xls_path.resolve())
        ]
        logger.info(f"Running LibreOffice conversion: {' '.join(cmd)}")
        result = subprocess.run(cmd, capture_output=True, text=True, timeout=60)
        
        if result.returncode == 0:
            expected_auto_name = xls_path.with_suffix('.xlsx')
            if expected_auto_name.exists() and expected_auto_name != xlsx_path:
                shutil.move(str(expected_auto_name), str(xlsx_path))
            return xlsx_path.exists()
        else:
            logger.error(f"Soffice conversion failed: {result.stderr}")
            return False
    except Exception as e:
        logger.error(f"Soffice error: {e}")
        return False

def _convert_xls_to_xlsx(folder_path: Path, excel_app=None):
    """Internal logic to convert all .xls files in a folder to .xlsx."""
    xls_files = glob.glob(str(folder_path / "*.xls"))
    if not xls_files:
        return

    try:
        for xls_path_str in xls_files:
            xls_path = Path(xls_path_str)
            if xls_path.suffix.lower() != ".xls": 
                continue
            
            xlsx_path = xls_path.with_suffix(".xlsx")
            if xlsx_path.exists() and xlsx_path.stat().st_mtime > xls_path.stat().st_mtime: 
                continue
            
            logger.info(f"Converting {xls_path.name} -> .xlsx")
            conversion_success = False

            # 1. Magic byte check
            try:
                with open(xls_path, 'rb') as f:
                    if f.read(2) == b'PK':
                        shutil.copy(xls_path, xlsx_path)
                        conversion_success = True
            except: pass

            # 2. Pure Python Strategy (xlrd -> openpyxl) - Most Stable
            if not conversion_success:
                try:
                    import xlrd
                    # xlrd 2.0+ only supports .xls, which is exactly what we want here
                    src_book = xlrd.open_workbook(str(xls_path), formatting_info=False)
                    out_wb = Workbook()
                    if "Sheet" in out_wb.sheetnames:
                        out_wb.remove(out_wb["Sheet"])

                    for i in range(src_book.nsheets):
                        src_ws = src_book.sheet_by_index(i)
                        # Clean sheet name for openpyxl
                        safe_name = re.sub(r'[\\/*?:\[\]]', '', (src_ws.name or f"Sheet{i+1}")).strip()[:31]
                        if not safe_name:
                            safe_name = f"Sheet{i+1}"
                        
                        # Handle duplicate names in the same workbook
                        orig_safe_name = safe_name
                        counter = 1
                        while safe_name in out_wb.sheetnames:
                            suffix = f"_{counter}"
                            safe_name = f"{orig_safe_name[:31-len(suffix)]}{suffix}"
                            counter += 1

                        out_ws = out_wb.create_sheet(title=safe_name)
                        for r in range(src_ws.nrows):
                            row_vals = src_ws.row_values(r)
                            for c, val in enumerate(row_vals, start=1):
                                write_val = val
                                if isinstance(write_val, str):
                                    write_val = ILLEGAL_CHARACTERS_RE.sub('', write_val)
                                    if write_val.startswith("="):
                                        write_val = "'" + write_val
                                out_ws.cell(row=r + 1, column=c, value=write_val)

                    out_wb.save(str(xlsx_path))
                    conversion_success = True
                    logger.debug(f"  Successfully converted {xls_path.name} using xlrd")
                except Exception as e:
                    logger.debug(f"  xlrd conversion failed: {e}")

            # 3. Windows COM Strategy (Fallback)
            if not conversion_success and os.name == 'nt':
                try:
                    import win32com.client
                    import pythoncom
                    abs_xls, abs_xlsx = str(xls_path.resolve()), str(xlsx_path.resolve())
                    
                    managed_excel = False
                    if excel_app is None:
                        pythoncom.CoInitialize()
                        # Use DispatchEx for isolation
                        excel = win32com.client.DispatchEx("Excel.Application")
                        excel.Visible = False
                        excel.DisplayAlerts = False
                        managed_excel = True
                    else:
                        excel = excel_app
                    
                    try:
                        wb = excel.Workbooks.Open(abs_xls, UpdateLinks=0)
                        wb.SaveAs(abs_xlsx, FileFormat=51)
                        wb.Close()
                        conversion_success = True
                        logger.debug(f"  Successfully converted {xls_path.name} using COM")
                    finally:
                        if managed_excel:
                            excel.Quit()
                            pythoncom.CoUninitialize()
                except Exception as e:
                    logger.debug(f"  COM Conversion error: {e}")

            # 4. Linux/macOS LibreOffice Strategy (High Fidelity Fallback)
            if not conversion_success and os.name != 'nt' and is_soffice_available():
                conversion_success = convert_xls_to_xlsx_soffice(xls_path, xlsx_path)

            if not conversion_success:
                logger.error(f"Failed to convert {xls_path.name} to .xlsx; source may require COM or LibreOffice conversion.")

    except Exception as e: 
        logger.error(f"XLS Conversion error: {e}")

def _get_clean_sheet_name(file_name: str, sheet_title: str, existing_names: List[str]) -> str:
    """Refactored sheet name cleaning logic."""
    decoded_file_name = unquote(file_name)
    base_raw_name = f"{decoded_file_name.rsplit('.', 1)[0]} {sheet_title}"
    sheet_name = base_raw_name.replace("_", " ").replace("-", " ").strip()
    
    days = [str(i).zfill(2) for i in range(1, 40)] + [str(i) for i in range(1, 40)]
    strip_keywords = [
        'ICICI', 'Prudential', 'HDFC', 'SBI', 'Axis', 'Nippon', 'India', 'Tata', 'Quant', 
        'Bandhan', 'Kotak', 'Aditya', 'Birla', 'Sun', 'Life', 'Mutual', 'Fund', 'MF', 'AMC',
        'Canara', 'Robeco', 'BNP', 'Paribas', 'Capitalmind',
        'Monthly', 'Portfolio', 'Disclosure', 'Statement',
        'January', 'February', 'March', 'April', 'May', 'June', 'July', 'August', 
        'September', 'October', 'November', 'December', '2024', '2025', '2026',
        'Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec',
    ] + days
    
    pattern = r'\b(' + '|'.join(set(strip_keywords)) + r')\b'
    sheet_name = re.sub(pattern, '', sheet_name, flags=re.IGNORECASE)
    sheet_name = re.sub(r'\b[0-9a-f]{4,}\b', '', sheet_name, flags=re.IGNORECASE)
    sheet_name = re.sub(r'\s+', ' ', sheet_name).strip()
    
    if not sheet_name or len(sheet_name) < 3:
        sheet_name = base_raw_name[:31]
        
    base_sheet_name = sheet_name[:31].strip()
    sheet_name = base_sheet_name
    
    counter = 1
    while sheet_name in existing_names:
        suffix = f" ({counter})"
        available_len = 31 - len(suffix)
        sheet_name = f"{base_sheet_name[:available_len].strip()}{suffix}"
        counter += 1
    return sheet_name

def merge_project_excels(folder_path: Path, output_filename: Path) -> Optional[Path]:
    """Smart entry point: prefers COM on Windows, fallbacks to openpyxl."""
    if os.name == 'nt':
        result = merge_project_excels_com(folder_path, output_filename)
        if result: return result
        logger.warning("COM merge failed or unavailable. Falling back to openpyxl merge.")
    
    return merge_project_excels_openpyxl(folder_path, output_filename)

def consolidate_amc_downloads(amc_slug: str, year: int, month: int) -> Optional[Path]:
    """
    Helper to consolidate all downloads for a specific AMC and period.
    Standardizes output to data/output/merged excels/{amc_slug}/{year}/
    """
    raw_folder = Path(f"data/raw/{amc_slug}/{year}_{month:02d}")
    if not raw_folder.exists():
        logger.warning(f"Raw folder {raw_folder} does not exist. Cannot consolidate.")
        return None
    
    output_folder = Path(f"data/output/merged excels/{amc_slug}/{year}")
    output_filename = output_folder / f"CONSOLIDATED_{amc_slug.upper()}_{year}_{month:02d}.xlsx"
    
    if output_filename.exists():
        try:
            output_mtime = output_filename.stat().st_mtime
            raw_files = list(raw_folder.glob("*.xls*"))
            if raw_files:
                latest_raw_mtime = max(f.stat().st_mtime for f in raw_files)
                if output_mtime > latest_raw_mtime:
                    logger.info(f"Consolidated file is up to date: {output_filename}")
                    notifier = get_notifier()
                    notifier.notify_merge_success(
                        amc=amc_slug.upper(), year=year, month=month, output_file=str(output_filename)
                    )
                    return output_filename
        except Exception as e:
            logger.warning(f"Error checking timestamps, forcing update: {e}")
    
    result = merge_project_excels(raw_folder, output_filename)
    
    notifier = get_notifier()
    if result:
        notifier.notify_merge_success(
            amc=amc_slug.upper(), year=year, month=month, output_file=str(result)
        )
    else:
        notifier.notify_merge_error(
            amc=amc_slug.upper(), year=year, month=month, error="Merging failed or no valid sheets found"
        )
            
    return result
